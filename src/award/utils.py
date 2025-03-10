import io
import logging
import random
import re
import string
from concurrent.futures import ProcessPoolExecutor
from functools import partial
from pathlib import Path
from pprint import pformat

import pymupdf
import pandas as pd
import pytesseract
import textract
from PIL import Image
from tqdm import tqdm
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


def process_file(file, keywords):
    try:
        text = extract_text_from_doc(file, totxt=False)
        return str(file), search_keywords_in_text(text, keywords)
    except Exception as e:
        logging.error("Error processing %s: %s", file, e)
        return None


def generate_random_string(length=6):
    characters = string.ascii_letters + string.digits
    return "".join(random.choice(characters) for _ in range(length))


def get_document_code_description(file_path: Path | str) -> tuple:
    file_path = Path(file_path)
    stem = file_path.stem
    match = re.search(r"[^\d_]", stem)
    if match:
        index = match.start() - 1 if match.start() > 1 else None
        code = stem[:index] if index else f"NoCode_{generate_random_string()}"
        description = stem[index + 1 :] if index is not None else stem
    else:
        code = stem
        description = None
    return code, description


def parse_code_and_description(df: pd.DataFrame) -> pd.DataFrame:
    leftmost_columns = ["Code", "Description"]
    df[leftmost_columns] = df["File"].apply(lambda x: pd.Series(get_document_code_description(x)))
    df = df.drop(columns=["File"])
    df = df[leftmost_columns + [col for col in df.columns if col not in leftmost_columns]]
    return df


def write_results_to_table(results: dict, output_file: str, format="xlsx") -> None:
    df = pd.DataFrame.from_dict(results, orient="index").reset_index()
    df = df.rename(columns={"index": "File"})
    df = parse_code_and_description(df)
    df["Score"] = df.iloc[:, 2:].sum(axis=1)

    # Insert the new column after the 'Description' column
    description_index = df.columns.get_loc("Description")
    df.insert(description_index + 1, "Score", df.pop("Score"))

    df = df.sort_values(by=["Score", "Code", "Description"], ascending=[False, True, True])
    if format == "csv":
        df.to_csv(output_file, index=False)
    elif format == "xlsx":
        df.to_excel(output_file, index=False)

    # Apply conditional formatting using openpyxl
    apply_conditional_formatting(output_file)


def apply_conditional_formatting(output_file: str) -> None:
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Find the column with the "Score" heading
    score_col_letter = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Score":
            score_col_letter = col[0].column_letter
            break

    if score_col_letter is None:
        raise ValueError("No 'Score' column found in the worksheet")

    # Define the color scale rule for the 'Score' column
    color_scale_rule = ColorScaleRule(
        start_type="min",
        start_color="FFFFFF",
        mid_type="percentile",
        mid_value=50,
        mid_color="FF9999",
        end_type="max",
        end_color="FF0000",
    )

    # Apply the color scale rule to the 'Score' column
    ws.conditional_formatting.add(
        f"{score_col_letter}2:{score_col_letter}{ws.max_row}", color_scale_rule
    )

    wb.save(output_file)


def get_keywords(file_path: Path) -> tuple:
    with open(file_path, "r") as file:
        lines = tuple(line.strip().lower() for line in file.readlines())
    return lines


def export_text_to_file(text: str, output_path: Path) -> None:
    with open(output_path, "w") as file:
        file.write(text)


def search_keywords_in_text(text: str, keywords: tuple) -> dict:
    results = {
        keyword: bool(re.search(rf"\b{re.escape(keyword)}\b", text, re.IGNORECASE))
        for keyword in keywords
    }
    return results


def extract_text_from_doc_simple(pdf_path: Path) -> tuple[str, int]:
    """Extract text content from a machine-readable PDF file.

    Args:
        pdf_path (Path): Path to the PDF file to process

    Returns:
        tuple: A tuple containing:
            - str: The extracted text from all pages concatenated
            - int: Total number of pages in the PDF
    """
    pdf_path = Path(pdf_path)
    document = pymupdf.open(pdf_path)
    page_count = document.page_count

    if page_count < 1:
        raise ValueError("No pages found. File may be corrupted or empty.")

    text = ""
    for page_num in range(page_count):
        page = document.load_page(page_num)
        page_text = page.get_text()
        text += page_text

    return text, document.page_count


def extract_text_from_pdf_ocr(pdf_path: Path) -> str:
    pdf_path = Path(pdf_path)
    text = ""
    document = pymupdf.open(pdf_path)
    for page_num in range(document.page_count):
        page = document.load_page(page_num)
        pix = page.get_pixmap()
        img = Image.open(io.BytesIO(pix.tobytes()))
        page_text = pytesseract.image_to_string(img)

        text += page_text

    return text


def extract_text_from_doc(file_path: Path, totxt=False) -> str:
    logging.info("Processing %s", file_path)
    extension = file_path.suffix.lower()
    # first try simple if document is machine readable
    if extension != ".doc":
        text, page_count = extract_text_from_doc_simple(file_path)
        logging.info("Extracted %d pages from %s", page_count, file_path)

        # if text is too short, try OCR
        if len(text.split("\n")) < 10 * page_count:
            logging.info("Warning: File is not machine readable. Trying OCR.")
            text = extract_text_from_pdf_ocr(file_path)
            # TODO: apply spellcheck to OCR text

    else:
        text = read_doc(file_path)
        logging.info("Succesfully processed %s", file_path)

    if totxt:
        out_name = file_path.stem
        out_name = Path(out_name).with_suffix(".txt")
        export_text_to_file(text, out_name)
        logging.info("Text extracted and saved to %s", out_name)

    return text


def read_doc(file_path):
    text = textract.process(file_path).decode("utf-8")
    return text


def find_files_with_extensions(
    input_dir: Path, extensions: tuple = (".pdf", ".docx", ".doc")
) -> list:
    """Find files with specified extensions in a directory.

    Args:
        input_dir (Path): Target directory to search for files.
        extensions (tuple, optional): Defaults to (".pdf", ".docx", ".doc").

    Returns:
        list: A list of file paths matching the specified extensions.
    """
    patterns = [f"**/*{ext}" for ext in extensions]
    return sorted([file for pattern in patterns for file in input_dir.glob(pattern)])


def search_files(
    input_dir,
    keywords_file,
    output_file,
) -> None:
    keywords = get_keywords(keywords_file)
    results = {}
    failures = []
    input_path = Path(input_dir)
    all_paths = find_files_with_extensions(input_path)

    process_file_with_keywords = partial(process_file, keywords=keywords)
    with ProcessPoolExecutor() as executor:
        for result in tqdm(
            executor.map(process_file_with_keywords, all_paths), total=len(all_paths)
        ):
            if result:
                file, keywords_result = result
                results[file] = keywords_result
            else:
                failures.append(file)

    if failures:
        logging.warning(
            "Failed to process %d files:\n%s",
            len(failures),
            pformat([str(f) for f in failures]),
        )

    if results:
        write_results_to_table(results, output_file)
